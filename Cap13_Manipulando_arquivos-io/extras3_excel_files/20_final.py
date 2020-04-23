#!/usr/bin/env python3

# template
# wb = Workbook

import openpyxl

wb = openpyxl.Workbook()  # carrega a planilha na memoria


wb.create_sheet(index=0, title='Horti-fruti')
wb.create_sheet(index=1, title='Grãos')
wb.create_sheet(index=2, title='Outros')

#obtendo os nomes das folhas(sheets)
#print(wb.sheetnames)

# inserindo conteudo nas folhas
sheet = wb['Horti-fruti']
sheet['A1'] = "Produto"
sheet['B1'] = "Quantidade"
sheet['C1'] = "Preço por Kg"
sheet['D1'] = "Quantidade em Estoque"
sheet['E1'] = "Vendido"
sheet['F1'] = "Total"
sheet['G1'] = "Data"


hf = {'produto':'Maçã', 'quantidade':None, 'preço_por_kg':None, 'qtd_eM_estoque':None,
      'vendidos':None, 'total':None, 'data':None}

lista_hf = ['maca']
sheet['A2'] = lista_hf[0]

# Obtendo valores
print(sheet['A1'].value)
print(sheet['A2'].value)
# print(sheet['B1'].value)
# print(sheet['C1'].value)
# print(sheet['D1'].value)
# print(sheet['E1'].value)
# print(sheet['F1'].value)
# print(sheet['G1'].value)
# print()


# Lendo colunas
# for i in range(1, sheet.max_row):
#     print(i, sheet.cell(row=i, column=1).value)
#
#
# # Salvando como excel(.xlsx)
# wb.save('feiradozero.xlsx')

