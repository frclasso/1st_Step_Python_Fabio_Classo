#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb['Sheet']
wb.create_sheet(index=1, title='Hortifruti')
wb.create_sheet(index=2, title='Grãos')
wb.create_sheet(index=3, title='Outros')

#print(wb.sheetnames)

# Obtendo valores
#print('Sheet')
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet['C1'].value)
print(sheet['D1'].value)
print()
# inserindo em 'Hortifrut'
#print('Hortifruti')
sheet_hf = wb['Hortifruti']

# sheet_hf['A1'] = sheet['A1'].value
# sheet_hf['B1'] = sheet['B1'].value
# sheet_hf['C1'] = sheet['C1'].value
# sheet_hf['D1'] = sheet['D1'].value

print(sheet_hf['A1'].value)
# print(sheet_hf['B1'].value)
# print(sheet_hf['C1'].value)
# print(sheet_hf['D1'].value)
# print()

# inserindo em 'Graos'
#print('Grãos')
sheet_g = wb['Grãos']
sheet_g['A1'] = sheet['A1'].value
sheet_g['B1'] = sheet['B1'].value
sheet_g['C1'] = sheet['C1'].value
sheet_g['D1'] = sheet['D1'].value

# print(sheet_g['A1'].value)
# print(sheet_g['B1'].value)
# print(sheet_g['C1'].value)
# print(sheet_g['D1'].value)
# print()

# inserindo em 'Outros'
#print('Outros')
sheet_o = wb['Outros']
sheet_o['A1'] = sheet['A1'].value
sheet_o['B1'] = sheet['B1'].value
sheet_o['C1'] = sheet['C1'].value
sheet_o['D1'] = sheet['D1'].value

# print(sheet_o['A1'].value)
# print(sheet_o['B1'].value)
# print(sheet_o['C1'].value)
# print(sheet_o['D1'].value)

# print('Qtd máxima de linhas: ', sheet.max_row)
# print('Qtd máxima de colunas: ', sheet.max_column)

# Lendo colunas
sheet_column_A = []
for i in range(1, sheet.max_row):
    #print(i, sheet.cell(row=i, column=1).value)
    sheet_column_A.append(sheet.cell(row=i, column=1).value)
#print(sheet_column_A)

# salvando
wb.save('feira_from_produce_sales.xlsx')
